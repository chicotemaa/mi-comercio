#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import os
import re
import statistics
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as error:  # pragma: no cover - runtime guidance
    raise SystemExit(
        "Falta openpyxl. Crea un venv e instala la dependencia:\n"
        "python3 -m venv .codex-import-venv\n"
        ".codex-import-venv/bin/pip install openpyxl\n"
        ".codex-import-venv/bin/python scripts/import_barber_excel.py --dry-run"
    ) from error


BASE_DATOS_SHEET = "BASE_DATOS"
EMPLEADOS_SHEET = "EMPLEADOS"
HORAS_SHEET = "HORAS"

DEFAULT_EXCEL_PATH = Path("/Users/matidev/Downloads/planilla_barberia.xlsx")
DEFAULT_REVIEW_PATH = Path(".codex-import-review.json")

SERVICE_ROW_TIME = "12:00:00"
PAYMENT_TIMESTAMP_SUFFIX = "T12:00:00-03:00"

PAYMENT_METHOD_MAP = {
    "EFECTIVO": "cash",
    "TRANSFERENCIA": "transfer",
    "TARJETA": "card",
    "MERCADO PAGO": "mercado_pago",
    "MP": "mercado_pago",
}

NON_CUSTOMER_KEYWORDS = {
    "ACEITE",
    "AGUA",
    "ALMUERZO",
    "ALQUILER",
    "CAMBIO",
    "CERA",
    "CEPILLO",
    "COCA",
    "CURSO",
    "EFECTIVO",
    "FABIAN",
    "FLAVIO",
    "GASTO",
    "GRAFICA",
    "HELADO",
    "INTERNET",
    "LUZ",
    "MAQUINA",
    "MERIENDA",
    "NAVAJA",
    "PRESTAMO",
    "SERVICIO TECNICO",
    "UBER",
    "WET LOOK",
    "ZORZON",
}

EXPENSE_CATEGORY_KEYWORDS = {
    "comida": {"ALMUERZO", "MERIENDA", "HELADO", "AGUA", "COCA"},
    "transporte": {"UBER"},
    "servicios": {"LUZ", "INTERNET"},
    "alquiler": {"ALQUILER"},
    "insumos": {"ACEITE", "CEPILLO", "WET LOOK", "NAVAJA", "CERA"},
}

PAYOUT_CATEGORY_KEYWORDS = {
    "staff_payment": {"AGOS", "PAGO AGOS", "PAGO SEMANAL AGOS", "PAGO SEMANA"},
    "supplier": {"FABIAN", "FLAVIO", "ZORZON", "GRAFICA", "CERA"},
}

SERVICE_CATEGORY_KEYWORDS = {
    "coloraciones": {"COLOR", "MECHAS", "BALAYAGE", "DECOLOR"},
    "tratamiento": {"NANO", "ALIS", "NUTRIC", "AMPOLLA", "PERMANENTE", "ELIXIR"},
}

SERVICE_DURATION_BY_CATEGORY = {
    "corte": 45,
    "coloraciones": 180,
    "tratamiento": 120,
}


@dataclass
class BaseDatosRow:
    row_number: int
    entry_date: date | None
    client_label: str
    service_name: str
    amount: float
    payment_method_raw: str
    distributor_amount: float
    expense_amount: float
    notes: str


@dataclass
class EmployeeSheetRow:
    row_number: int
    employee_code: str
    full_name: str
    role: str
    hourly_rate: float
    is_active: bool


@dataclass
class HourSheetRow:
    row_number: int
    work_date: date | None
    employee_code: str
    start_time: time | None
    end_time: time | None
    hours_worked: float
    entry_type: str
    notes: str


@dataclass
class CustomerState:
    id: str
    full_name: str
    primary_contact: str
    total_spent: float
    total_appointments: int
    joined_at: str
    last_visit_at: str | None
    preferred_services: list[str]
    preferred_service_counts: dict[str, int] = field(default_factory=dict)
    dirty: bool = False


@dataclass
class ServiceState:
    id: str
    name: str
    duration_minutes: int
    has_default_variant: bool


@dataclass
class ImportSummary:
    services_created: int = 0
    services_skipped: int = 0
    staff_created: int = 0
    staff_updated: int = 0
    staff_skipped: int = 0
    customers_created: int = 0
    customers_updated: int = 0
    appointments_created: int = 0
    appointments_skipped: int = 0
    payments_created: int = 0
    payments_skipped: int = 0
    expenses_created: int = 0
    expenses_skipped: int = 0
    payouts_created: int = 0
    payouts_skipped: int = 0
    time_logs_created: int = 0
    time_logs_skipped: int = 0


def collapse_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_key(value: str) -> str:
    collapsed = collapse_spaces(value)
    normalized = unicodedata.normalize("NFKD", collapsed)
    without_marks = "".join(char for char in normalized if not unicodedata.combining(char))
    uppercased = without_marks.upper()
    sanitized = re.sub(r"[^A-Z0-9 ]+", " ", uppercased)
    return collapse_spaces(sanitized)


def strip_trailing_markers(value: str) -> str:
    trimmed = collapse_spaces(value)
    return collapse_spaces(
        re.sub(
            r"(?i)\b(?:x\d+|\d+|mensual|mes|saldo|sald[oó]|seña|seño)\b$",
            "",
            trimmed,
        )
    )


def customer_key(value: str) -> str:
    return normalize_key(strip_trailing_markers(value))


def prettify_name(value: str) -> str:
    cleaned = strip_trailing_markers(value)
    return collapse_spaces(cleaned).title()


def normalize_service_name(value: str) -> str:
    return collapse_spaces(value).upper()


def coerce_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        parsed = datetime.fromisoformat(value.strip())
        return parsed.date()
    return None


def coerce_time(value: Any) -> time | None:
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, str) and value.strip():
        text = value.strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(text, fmt).time()
            except ValueError:
                continue
    return None


def coerce_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "si", "sí", "yes"}
    return bool(value)


def load_local_env() -> None:
    for path in (Path(".env.local"), Path(".env")):
        if not path.exists():
            continue
        for raw_line in path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip().strip("'").strip('"')
            os.environ.setdefault(key, value)


def build_query_string(params: dict[str, Any]) -> str:
    parts: list[str] = []
    for key, value in params.items():
        encoded_key = urllib.parse.quote(str(key), safe="")
        encoded_value = urllib.parse.quote(str(value), safe="(),.*:-_")
        parts.append(f"{encoded_key}={encoded_value}")
    return "&".join(parts)


class SupabaseRestClient:
    def __init__(self, base_url: str, service_key: str):
        self.base_url = base_url.rstrip("/")
        self.service_key = service_key
        self.api_base = f"{self.base_url}/rest/v1"

    def request(
        self,
        method: str,
        path: str,
        *,
        params: dict[str, Any] | None = None,
        body: Any | None = None,
        headers: dict[str, str] | None = None,
    ) -> Any:
        url = f"{self.api_base}/{path.lstrip('/')}"
        if params:
            url = f"{url}?{build_query_string(params)}"

        request_headers = {
            "apikey": self.service_key,
            "Authorization": f"Bearer {self.service_key}",
            "Accept": "application/json",
            "Range": "0-9999",
        }
        if headers:
            request_headers.update(headers)

        payload = None
        if body is not None:
            payload = json.dumps(body).encode("utf-8")
            request_headers["Content-Type"] = "application/json"

        request = urllib.request.Request(
            url,
            data=payload,
            method=method.upper(),
            headers=request_headers,
        )

        try:
            with urllib.request.urlopen(request) as response:
                raw_response = response.read().decode("utf-8")
        except urllib.error.HTTPError as error:
            message = error.read().decode("utf-8")
            raise RuntimeError(
                f"Supabase {method.upper()} {path} devolvió {error.code}: {message}"
            ) from error

        if not raw_response:
            return None

        return json.loads(raw_response)

    def select(self, table: str, params: dict[str, Any]) -> list[dict[str, Any]]:
        response = self.request("GET", table, params=params)
        return response if isinstance(response, list) else []

    def insert(
        self,
        table: str,
        rows: list[dict[str, Any]] | dict[str, Any],
        *,
        select: str = "*",
    ) -> list[dict[str, Any]]:
        response = self.request(
            "POST",
            table,
            params={"select": select},
            body=rows,
            headers={"Prefer": "return=representation"},
        )
        if response is None:
            return []
        return response if isinstance(response, list) else [response]

    def patch(
        self,
        table: str,
        filters: dict[str, Any],
        values: dict[str, Any],
        *,
        select: str = "*",
    ) -> list[dict[str, Any]]:
        params = {"select": select}
        params.update(filters)
        response = self.request(
            "PATCH",
            table,
            params=params,
            body=values,
            headers={"Prefer": "return=representation"},
        )
        if response is None:
            return []
        return response if isinstance(response, list) else [response]


def get_required_env(name: str) -> str:
    value = os.environ.get(name)
    if not value:
        raise SystemExit(f"Falta la variable {name} en .env.local o en el entorno.")
    return value


def read_base_datos(path: Path) -> list[BaseDatosRow]:
    workbook = load_workbook(path, data_only=True)
    if BASE_DATOS_SHEET not in workbook.sheetnames:
        raise SystemExit(f"No existe la hoja {BASE_DATOS_SHEET} en {path}.")

    rows: list[BaseDatosRow] = []
    sheet = workbook[BASE_DATOS_SHEET]
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        entry = BaseDatosRow(
            row_number=row_number,
            entry_date=coerce_date(values[0]),
            client_label=collapse_spaces(str(values[1] or "")),
            service_name=collapse_spaces(str(values[2] or "")),
            amount=coerce_float(values[3]),
            payment_method_raw=collapse_spaces(str(values[4] or "")).upper(),
            distributor_amount=coerce_float(values[5]),
            expense_amount=coerce_float(values[6]),
            notes=collapse_spaces(str(values[7] or "")),
        )
        if any(
            [
                entry.entry_date,
                entry.client_label,
                entry.service_name,
                entry.amount,
                entry.payment_method_raw,
                entry.distributor_amount,
                entry.expense_amount,
                entry.notes,
            ]
        ):
            rows.append(entry)
    return rows


def read_employee_sheet(path: Path) -> list[EmployeeSheetRow]:
    workbook = load_workbook(path, data_only=True)
    if EMPLEADOS_SHEET not in workbook.sheetnames:
        return []

    rows: list[EmployeeSheetRow] = []
    sheet = workbook[EMPLEADOS_SHEET]
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        employee_code = collapse_spaces(str(values[0] or "")).upper()
        full_name = collapse_spaces(str(values[1] or ""))
        if not employee_code and not full_name:
            continue
        rows.append(
            EmployeeSheetRow(
                row_number=row_number,
                employee_code=employee_code,
                full_name=prettify_name(full_name or employee_code),
                role=collapse_spaces(str(values[2] or "")),
                hourly_rate=coerce_float(values[3]),
                is_active=parse_bool(values[4]),
            )
        )
    return rows


def read_hours_sheet(path: Path) -> list[HourSheetRow]:
    workbook = load_workbook(path, data_only=True)
    if HORAS_SHEET not in workbook.sheetnames:
        return []

    rows: list[HourSheetRow] = []
    sheet = workbook[HORAS_SHEET]
    for row_number, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        work_date = coerce_date(values[0])
        employee_code = collapse_spaces(str(values[1] or "")).upper()
        if not work_date and not employee_code:
            continue
        rows.append(
            HourSheetRow(
                row_number=row_number,
                work_date=work_date,
                employee_code=employee_code,
                start_time=coerce_time(values[2]),
                end_time=coerce_time(values[3]),
                hours_worked=coerce_float(values[4]),
                entry_type=collapse_spaces(str(values[5] or "")) or "shift",
                notes=collapse_spaces(str(values[6] or "")),
            )
        )
    return rows


def infer_payment_method(raw_value: str, default: str) -> str:
    normalized = normalize_key(raw_value)
    return PAYMENT_METHOD_MAP.get(normalized, default)


def contains_keyword(text: str, candidates: set[str]) -> bool:
    normalized = normalize_key(text)
    return any(candidate in normalized for candidate in candidates)


def is_customer_like(label: str) -> bool:
    if not label:
        return False
    normalized = customer_key(label)
    if not normalized:
        return False
    return not any(keyword in normalized for keyword in NON_CUSTOMER_KEYWORDS)


def infer_expense_category(label: str, notes: str) -> str:
    combined = f"{label} {notes}"
    for category, keywords in EXPENSE_CATEGORY_KEYWORDS.items():
        if contains_keyword(combined, keywords):
            return category
    return "operativo"


def infer_payout_category(label: str, notes: str) -> tuple[str, str]:
    combined = f"{label} {notes}"
    for category, keywords in PAYOUT_CATEGORY_KEYWORDS.items():
        if contains_keyword(combined, keywords):
            recipient_type = "staff" if category == "staff_payment" else "vendor"
            return category, recipient_type
    return "distribution", "vendor"


def infer_service_category(service_name: str) -> str:
    normalized = normalize_key(service_name)
    for category, keywords in SERVICE_CATEGORY_KEYWORDS.items():
        if any(keyword in normalized for keyword in keywords):
            return category
    return "corte"


def infer_service_duration(service_name: str, category: str) -> int:
    normalized = normalize_key(service_name)
    if "CORTE Y BARBA" in normalized:
        return 60
    if "BARBA" in normalized or "CEJAS" in normalized or "PERFILADO" in normalized:
        return 30
    if "MUJER" in normalized:
        return 60
    return SERVICE_DURATION_BY_CATEGORY[category]


def infer_service_description(service_name: str, category: str) -> str:
    label = normalize_service_name(service_name).title()
    if category == "coloraciones":
        return f"{label} importado desde la planilla histórica del negocio."
    if category == "tratamiento":
        return f"{label} importado como tratamiento histórico desde Excel."
    return f"{label} importado como servicio histórico desde Excel."


def ensure_business(client: SupabaseRestClient, business_slug: str) -> dict[str, Any]:
    rows = client.select(
        "businesses",
        {
            "select": "id,name,slug",
            "slug": f"eq.{business_slug}",
        },
    )
    if len(rows) != 1:
        raise SystemExit(
            f"No se pudo resolver el negocio con BUSINESS_SLUG={business_slug}."
        )
    return rows[0]


def fetch_services(
    client: SupabaseRestClient,
    business_id: str,
) -> dict[str, ServiceState]:
    rows = client.select(
        "services",
        {
            "select": "id,name,duration_minutes,display_order",
            "business_id": f"eq.{business_id}",
        },
    )
    default_variant_rows = client.select(
        "service_price_variants",
        {
            "select": "service_id,is_default",
            "is_default": "eq.true",
        },
    )
    service_variant_ids = {
        row["service_id"]
        for row in default_variant_rows
        if row.get("service_id")
    }

    services: dict[str, ServiceState] = {}
    for row in rows:
        name = str(row.get("name") or "")
        key = normalize_key(name)
        if not key:
            continue
        services[key] = ServiceState(
            id=row["id"],
            name=name,
            duration_minutes=int(row.get("duration_minutes") or 45),
            has_default_variant=row["id"] in service_variant_ids,
        )
    return services


def fetch_staff(
    client: SupabaseRestClient,
    business_id: str,
) -> dict[str, dict[str, Any]]:
    rows = client.select(
        "staff_members",
        {
            "select": "id,full_name,employee_code,hourly_rate,role,is_active",
            "business_id": f"eq.{business_id}",
        },
    )
    staff: dict[str, dict[str, Any]] = {}
    for row in rows:
        full_name = str(row.get("full_name") or "")
        employee_code = str(row.get("employee_code") or "")
        if full_name:
            staff[f"name:{customer_key(full_name)}"] = row
        if employee_code:
            staff[f"code:{normalize_key(employee_code)}"] = row
    return staff


def fetch_customers(
    client: SupabaseRestClient,
    business_id: str,
) -> dict[str, CustomerState]:
    rows = client.select(
        "customers",
        {
            "select": "id,full_name,primary_contact,total_spent,total_appointments,last_visit_at,joined_at,preferred_services",
            "business_id": f"eq.{business_id}",
        },
    )
    customers: dict[str, CustomerState] = {}
    for row in rows:
        full_name = str(row.get("full_name") or "")
        key = customer_key(full_name)
        if not key:
            continue
        preferred_services = row.get("preferred_services") or []
        customers[key] = CustomerState(
            id=row["id"],
            full_name=full_name,
            primary_contact=str(row.get("primary_contact") or full_name),
            total_spent=coerce_float(row.get("total_spent")),
            total_appointments=int(row.get("total_appointments") or 0),
            joined_at=str(row.get("joined_at") or datetime.now().isoformat()),
            last_visit_at=row.get("last_visit_at"),
            preferred_services=list(preferred_services),
            preferred_service_counts={name: 1 for name in preferred_services},
        )
    return customers


def fetch_existing_markers(
    client: SupabaseRestClient,
    business_id: str,
    table: str,
    column: str,
) -> set[str]:
    rows = client.select(
        table,
        {
            "select": column,
            "business_id": f"eq.{business_id}",
            column: "not.is.null",
        },
    )
    values = set()
    for row in rows:
        value = row.get(column)
        if isinstance(value, str) and value:
            values.add(value)
    return values


def row_marker(sheet_name: str, row_number: int) -> str:
    return f"excel:{sheet_name}:{row_number}"


def iso_date(value: date) -> str:
    return value.isoformat()


def iso_datetime_for_payment(value: date) -> str:
    return f"{value.isoformat()}{PAYMENT_TIMESTAMP_SUFFIX}"


def time_to_string(value: time | None, fallback: str | None = None) -> str | None:
    if value is None:
        return fallback
    return value.strftime("%H:%M:%S")


def median_amount(amounts: list[float]) -> float:
    if not amounts:
        return 0.0
    return round(float(statistics.median(amounts)), 2)


def maybe_append_note(base_note: str, extra_note: str) -> str | None:
    pieces = [piece for piece in [base_note, extra_note] if piece]
    return " | ".join(pieces) if pieces else None


def sync_services(
    client: SupabaseRestClient,
    *,
    apply_changes: bool,
    business_id: str,
    base_rows: list[BaseDatosRow],
    services: dict[str, ServiceState],
    summary: ImportSummary,
) -> None:
    max_display_order = len(services)
    pending_variants: list[dict[str, Any]] = []
    service_observations: dict[str, list[float]] = {}

    for row in base_rows:
        if not row.service_name:
            continue
        key = normalize_key(row.service_name)
        service_observations.setdefault(key, [])
        if row.amount > 0:
            service_observations[key].append(row.amount)

    for row in base_rows:
        if not row.service_name:
            continue
        key = normalize_key(row.service_name)
        if key in services:
            summary.services_skipped += 1
            continue

        category = infer_service_category(row.service_name)
        duration_minutes = infer_service_duration(row.service_name, category)
        price = median_amount(service_observations.get(key, []))
        payload = {
            "business_id": business_id,
            "name": normalize_service_name(row.service_name),
            "description": infer_service_description(row.service_name, category),
            "duration_minutes": duration_minutes,
            "price": price,
            "is_active": True,
            "display_order": max_display_order,
            "category": category,
            "booking_enabled": True,
        }
        max_display_order += 1

        if apply_changes:
            created_rows = client.insert("services", payload, select="id,name,duration_minutes")
            if not created_rows:
                raise RuntimeError(f"No se pudo crear el servicio {row.service_name}.")
            created_service = created_rows[0]
            service_id = created_service["id"]
            client.insert(
                "service_price_variants",
                {
                    "service_id": service_id,
                    "variant_name": "Base",
                    "variant_code": "base",
                    "price": price,
                    "duration_minutes": duration_minutes,
                    "is_default": True,
                    "is_active": True,
                    "display_order": 1,
                    "notes": "Variante base importada desde Excel.",
                },
                select="id",
            )
            services[key] = ServiceState(
                id=service_id,
                name=created_service["name"],
                duration_minutes=int(created_service["duration_minutes"]),
                has_default_variant=True,
            )
        else:
            pending_variants.append(payload)
            services[key] = ServiceState(
                id=f"dry-run-{key}",
                name=payload["name"],
                duration_minutes=duration_minutes,
                has_default_variant=True,
            )
        summary.services_created += 1


def sync_staff(
    client: SupabaseRestClient,
    *,
    apply_changes: bool,
    business_id: str,
    employee_rows: list[EmployeeSheetRow],
    staff_lookup: dict[str, dict[str, Any]],
    review_items: list[dict[str, Any]],
    summary: ImportSummary,
) -> None:
    for row in employee_rows:
        by_code = staff_lookup.get(f"code:{normalize_key(row.employee_code)}")
        by_name = staff_lookup.get(f"name:{customer_key(row.full_name)}")
        matched = by_code or by_name

        payload = {
            "employee_code": row.employee_code or None,
            "full_name": row.full_name,
            "role": row.role or None,
            "hourly_rate": row.hourly_rate,
            "is_active": row.is_active,
        }

        if matched:
            needs_update = any(
                [
                    collapse_spaces(str(matched.get("full_name") or "")) != row.full_name,
                    collapse_spaces(str(matched.get("employee_code") or "")).upper()
                    != row.employee_code,
                    collapse_spaces(str(matched.get("role") or "")) != row.role,
                    round(coerce_float(matched.get("hourly_rate")), 2)
                    != round(row.hourly_rate, 2),
                    bool(matched.get("is_active")) != row.is_active,
                ]
            )

            if not needs_update:
                summary.staff_skipped += 1
                continue

            if apply_changes:
                client.patch(
                    "staff_members",
                    {"id": f"eq.{matched['id']}"},
                    payload,
                    select="id,full_name,employee_code,hourly_rate,role,is_active",
                )
            matched.update(payload)
            staff_lookup[f"name:{customer_key(row.full_name)}"] = matched
            if row.employee_code:
                staff_lookup[f"code:{normalize_key(row.employee_code)}"] = matched
            summary.staff_updated += 1
            continue

        if apply_changes:
            created_rows = client.insert(
                "staff_members",
                {
                    "business_id": business_id,
                    **payload,
                },
                select="id,full_name,employee_code,hourly_rate,role,is_active",
            )
            created = created_rows[0]
        else:
            created = {
                "id": f"dry-run-staff-{row.row_number}",
                **payload,
            }
        staff_lookup[f"name:{customer_key(row.full_name)}"] = created
        if row.employee_code:
            staff_lookup[f"code:{normalize_key(row.employee_code)}"] = created
        summary.staff_created += 1

        if not row.employee_code:
            review_items.append(
                {
                    "type": "staff_without_code",
                    "sheet": EMPLEADOS_SHEET,
                    "row": row.row_number,
                    "name": row.full_name,
                }
            )


def ensure_customer(
    client: SupabaseRestClient,
    *,
    apply_changes: bool,
    business_id: str,
    customers: dict[str, CustomerState],
    raw_label: str,
    entry_date: date | None,
    summary: ImportSummary,
) -> CustomerState | None:
    if not is_customer_like(raw_label):
        return None

    key = customer_key(raw_label)
    if not key:
        return None

    existing = customers.get(key)
    if existing:
        return existing

    pretty_name = prettify_name(raw_label)
    joined_at = (
        iso_datetime_for_payment(entry_date)
        if entry_date is not None
        else datetime.now().isoformat()
    )
    payload = {
        "business_id": business_id,
        "full_name": pretty_name,
        "primary_contact": pretty_name,
        "status": "active",
        "rating": 5,
        "marketing_opt_in": False,
        "joined_at": joined_at,
        "notes": "Cliente importado desde planilla_barberia.xlsx.",
    }

    if apply_changes:
        created_rows = client.insert(
            "customers",
            payload,
            select="id,full_name,primary_contact,total_spent,total_appointments,last_visit_at,joined_at,preferred_services",
        )
        created = created_rows[0]
        customer_state = CustomerState(
            id=created["id"],
            full_name=created["full_name"],
            primary_contact=created["primary_contact"],
            total_spent=coerce_float(created.get("total_spent")),
            total_appointments=int(created.get("total_appointments") or 0),
            joined_at=str(created.get("joined_at") or joined_at),
            last_visit_at=created.get("last_visit_at"),
            preferred_services=list(created.get("preferred_services") or []),
        )
    else:
        customer_state = CustomerState(
            id=f"dry-run-customer-{key}",
            full_name=pretty_name,
            primary_contact=pretty_name,
            total_spent=0.0,
            total_appointments=0,
            joined_at=joined_at,
            last_visit_at=None,
            preferred_services=[],
        )

    customers[key] = customer_state
    summary.customers_created += 1
    return customer_state


def update_customer_stats(
    customer: CustomerState,
    *,
    amount_delta: float = 0.0,
    appointment_delta: int = 0,
    visit_date: date | None = None,
    service_name: str | None = None,
) -> None:
    if amount_delta:
        customer.total_spent = round(customer.total_spent + amount_delta, 2)
        customer.dirty = True
    if appointment_delta:
        customer.total_appointments += appointment_delta
        customer.dirty = True
    if visit_date:
        visit_date_value = iso_date(visit_date)
        if not customer.last_visit_at or visit_date_value > customer.last_visit_at:
            customer.last_visit_at = visit_date_value
            customer.dirty = True
        joined_date = customer.joined_at[:10]
        if joined_date > visit_date_value:
            customer.joined_at = iso_datetime_for_payment(visit_date)
            customer.dirty = True
    if service_name:
        customer.preferred_service_counts[service_name] = (
            customer.preferred_service_counts.get(service_name, 0) + 1
        )
        customer.dirty = True


def flush_customer_updates(
    client: SupabaseRestClient,
    *,
    apply_changes: bool,
    customers: dict[str, CustomerState],
    summary: ImportSummary,
) -> None:
    for customer in customers.values():
        if not customer.dirty:
            continue
        preferred_services = [
            name
            for name, _count in sorted(
                customer.preferred_service_counts.items(),
                key=lambda item: (-item[1], item[0]),
            )[:5]
        ]
        payload = {
            "total_spent": round(customer.total_spent, 2),
            "total_appointments": customer.total_appointments,
            "last_visit_at": customer.last_visit_at,
            "joined_at": customer.joined_at,
            "preferred_services": preferred_services,
        }
        if apply_changes:
            client.patch("customers", {"id": f"eq.{customer.id}"}, payload, select="id")
        customer.preferred_services = preferred_services
        customer.dirty = False
        summary.customers_updated += 1


def import_base_rows(
    client: SupabaseRestClient,
    *,
    apply_changes: bool,
    business_id: str,
    base_rows: list[BaseDatosRow],
    services: dict[str, ServiceState],
    customers: dict[str, CustomerState],
    staff_lookup: dict[str, dict[str, Any]],
    appointment_markers: set[str],
    payment_markers: set[str],
    expense_markers: set[str],
    payout_markers: set[str],
    review_items: list[dict[str, Any]],
    summary: ImportSummary,
) -> None:
    appointment_ids_by_marker: dict[str, str] = {}

    for row in base_rows:
        if row.entry_date is None:
            review_items.append(
                {
                    "type": "missing_date",
                    "sheet": BASE_DATOS_SHEET,
                    "row": row.row_number,
                    "client_label": row.client_label,
                }
            )
            continue

        marker = row_marker(BASE_DATOS_SHEET, row.row_number)
        customer = ensure_customer(
            client,
            apply_changes=apply_changes,
            business_id=business_id,
            customers=customers,
            raw_label=row.client_label,
            entry_date=row.entry_date,
            summary=summary,
        )
        service_state = services.get(normalize_key(row.service_name)) if row.service_name else None

        if row.service_name:
            if marker in appointment_markers:
                summary.appointments_skipped += 1
            else:
                appointment_payload = {
                    "business_id": business_id,
                    "service_id": service_state.id if service_state else None,
                    "staff_member_id": None,
                    "customer_id": customer.id if customer else None,
                    "customer_name": prettify_name(row.client_label) if row.client_label else "Cliente sin nombre",
                    "customer_contact": prettify_name(row.client_label) if row.client_label else "Sin contacto",
                    "customer_email": None,
                    "appointment_date": iso_date(row.entry_date),
                    "appointment_time": SERVICE_ROW_TIME,
                    "status": "completed",
                    "channel": "manual",
                    "service_name_snapshot": normalize_service_name(row.service_name),
                    "staff_name_snapshot": None,
                    "price_snapshot": row.amount if row.amount > 0 else 0,
                    "duration_snapshot": service_state.duration_minutes if service_state else infer_service_duration(row.service_name, infer_service_category(row.service_name)),
                    "notes": maybe_append_note(row.notes, f"Importado desde Excel fila {row.row_number}."),
                    "internal_notes": marker,
                }
                if apply_changes:
                    created_rows = client.insert("appointments", appointment_payload, select="id")
                    appointment_ids_by_marker[marker] = created_rows[0]["id"]
                else:
                    appointment_ids_by_marker[marker] = f"dry-run-appointment-{row.row_number}"
                appointment_markers.add(marker)
                summary.appointments_created += 1
                if customer:
                    update_customer_stats(
                        customer,
                        appointment_delta=1,
                        visit_date=row.entry_date,
                        service_name=normalize_service_name(row.service_name),
                    )

            if row.amount <= 0:
                review_items.append(
                    {
                        "type": "service_without_positive_payment",
                        "sheet": BASE_DATOS_SHEET,
                        "row": row.row_number,
                        "client_label": row.client_label,
                        "service_name": row.service_name,
                        "notes": row.notes,
                    }
                )

        if row.amount > 0:
            if marker in payment_markers:
                summary.payments_skipped += 1
            else:
                payment_notes = row.notes
                customer_linked = customer is not None
                if not customer_linked and row.client_label:
                    payment_notes = maybe_append_note(
                        payment_notes,
                        f"Detalle original: {row.client_label}.",
                    ) or ""

                payment_description = (
                    normalize_service_name(row.service_name)
                    if row.service_name
                    else (row.notes or row.client_label or "Ingreso importado")
                )
                payment_payload = {
                    "business_id": business_id,
                    "customer_id": customer.id if customer else None,
                    "appointment_id": appointment_ids_by_marker.get(marker),
                    "staff_member_id": None,
                    "description": payment_description,
                    "amount": row.amount,
                    "method": infer_payment_method(row.payment_method_raw, "other"),
                    "status": "completed",
                    "transaction_id": marker,
                    "processed_at": iso_datetime_for_payment(row.entry_date),
                    "notes": maybe_append_note(payment_notes, f"Importado desde Excel fila {row.row_number}."),
                }
                if apply_changes:
                    client.insert("payments", payment_payload, select="id")
                payment_markers.add(marker)
                summary.payments_created += 1
                if customer:
                    update_customer_stats(
                        customer,
                        amount_delta=row.amount,
                        visit_date=row.entry_date,
                    )

            if row.notes and contains_keyword(row.notes, {"PRESTAMO", "CURSO"}):
                review_items.append(
                    {
                        "type": "ambiguous_income_note",
                        "sheet": BASE_DATOS_SHEET,
                        "row": row.row_number,
                        "client_label": row.client_label,
                        "amount": row.amount,
                        "notes": row.notes,
                    }
                )

        if row.expense_amount > 0:
            if marker in expense_markers:
                summary.expenses_skipped += 1
            else:
                expense_payload = {
                    "business_id": business_id,
                    "expense_date": iso_date(row.entry_date),
                    "category": infer_expense_category(row.client_label, row.notes),
                    "subcategory": row.client_label or None,
                    "description": row.client_label or row.notes or "Gasto importado",
                    "vendor_name": row.client_label or None,
                    "amount": row.expense_amount,
                    "method": infer_payment_method(row.payment_method_raw, "cash"),
                    "source": marker,
                    "notes": maybe_append_note(row.notes, f"Importado desde Excel fila {row.row_number}."),
                }
                if apply_changes:
                    client.insert("expenses", expense_payload, select="id")
                expense_markers.add(marker)
                summary.expenses_created += 1

        if row.distributor_amount > 0:
            if marker in payout_markers:
                summary.payouts_skipped += 1
            else:
                payout_category, recipient_type = infer_payout_category(
                    row.client_label,
                    row.notes,
                )
                staff_match = staff_lookup.get(f"code:{normalize_key(row.client_label)}") or staff_lookup.get(
                    f"name:{customer_key(row.client_label)}"
                )
                payout_payload = {
                    "business_id": business_id,
                    "payout_date": iso_date(row.entry_date),
                    "recipient_name": prettify_name(row.client_label) if row.client_label else "Destinatario sin nombre",
                    "recipient_type": "staff" if staff_match else recipient_type,
                    "category": payout_category,
                    "staff_member_id": staff_match["id"] if staff_match else None,
                    "amount": row.distributor_amount,
                    "method": infer_payment_method(row.payment_method_raw, "transfer"),
                    "source": marker,
                    "notes": maybe_append_note(row.notes, f"Importado desde Excel fila {row.row_number}."),
                }
                if apply_changes:
                    client.insert("payouts", payout_payload, select="id")
                payout_markers.add(marker)
                summary.payouts_created += 1

        if not any(
            [row.service_name, row.amount, row.distributor_amount, row.expense_amount]
        ):
            review_items.append(
                {
                    "type": "empty_business_row",
                    "sheet": BASE_DATOS_SHEET,
                    "row": row.row_number,
                    "client_label": row.client_label,
                    "notes": row.notes,
                }
            )


def import_hour_rows(
    client: SupabaseRestClient,
    *,
    apply_changes: bool,
    business_id: str,
    hour_rows: list[HourSheetRow],
    staff_lookup: dict[str, dict[str, Any]],
    time_log_markers: set[str],
    review_items: list[dict[str, Any]],
    summary: ImportSummary,
) -> None:
    for row in hour_rows:
        if row.work_date is None:
            review_items.append(
                {
                    "type": "hour_row_without_date",
                    "sheet": HORAS_SHEET,
                    "row": row.row_number,
                    "employee_code": row.employee_code,
                }
            )
            continue

        marker = row_marker(HORAS_SHEET, row.row_number)
        if marker in time_log_markers:
            summary.time_logs_skipped += 1
            continue

        staff_match = staff_lookup.get(f"code:{normalize_key(row.employee_code)}") or staff_lookup.get(
            f"name:{customer_key(row.employee_code)}"
        )
        if not staff_match:
            review_items.append(
                {
                    "type": "unknown_staff_time_log",
                    "sheet": HORAS_SHEET,
                    "row": row.row_number,
                    "employee_code": row.employee_code,
                    "hours_worked": row.hours_worked,
                }
            )
            continue

        payload = {
            "business_id": business_id,
            "staff_member_id": staff_match["id"],
            "work_date": iso_date(row.work_date),
            "start_time": time_to_string(row.start_time),
            "end_time": time_to_string(row.end_time),
            "hours_worked": row.hours_worked,
            "entry_type": row.entry_type or "shift",
            "source": marker,
            "notes": maybe_append_note(row.notes, f"Importado desde Excel fila {row.row_number}."),
        }

        if apply_changes:
            client.insert("staff_time_logs", payload, select="id")
        time_log_markers.add(marker)
        summary.time_logs_created += 1


def build_review_report(
    *,
    excel_path: Path,
    business_slug: str,
    apply_changes: bool,
    summary: ImportSummary,
    review_items: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "excelPath": str(excel_path),
        "businessSlug": business_slug,
        "mode": "apply" if apply_changes else "dry-run",
        "summary": {
            "servicesCreated": summary.services_created,
            "servicesSkipped": summary.services_skipped,
            "staffCreated": summary.staff_created,
            "staffUpdated": summary.staff_updated,
            "staffSkipped": summary.staff_skipped,
            "customersCreated": summary.customers_created,
            "customersUpdated": summary.customers_updated,
            "appointmentsCreated": summary.appointments_created,
            "appointmentsSkipped": summary.appointments_skipped,
            "paymentsCreated": summary.payments_created,
            "paymentsSkipped": summary.payments_skipped,
            "expensesCreated": summary.expenses_created,
            "expensesSkipped": summary.expenses_skipped,
            "payoutsCreated": summary.payouts_created,
            "payoutsSkipped": summary.payouts_skipped,
            "timeLogsCreated": summary.time_logs_created,
            "timeLogsSkipped": summary.time_logs_skipped,
            "reviewItems": len(review_items),
        },
        "reviewItems": review_items,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Importa la planilla histórica de barbería a Supabase.",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL_PATH,
        help=f"Ruta al Excel. Por defecto usa {DEFAULT_EXCEL_PATH}.",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Aplica cambios en Supabase. Sin esta bandera corre en dry-run.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Fuerza modo simulación aunque se haya dejado preparado el entorno.",
    )
    parser.add_argument(
        "--review-file",
        type=Path,
        default=DEFAULT_REVIEW_PATH,
        help=f"Ruta del reporte JSON. Por defecto usa {DEFAULT_REVIEW_PATH}.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.apply and args.dry_run:
        raise SystemExit("Usa solo uno entre --apply y --dry-run.")
    excel_path = args.excel.expanduser()
    if not excel_path.exists():
        raise SystemExit(f"No se encontró el Excel en {excel_path}.")

    load_local_env()
    supabase_url = get_required_env("NEXT_PUBLIC_SUPABASE_URL")
    supabase_service_role_key = get_required_env("SUPABASE_SERVICE_ROLE_KEY")
    business_slug = get_required_env("BUSINESS_SLUG")

    client = SupabaseRestClient(supabase_url, supabase_service_role_key)
    business = ensure_business(client, business_slug)

    base_rows = read_base_datos(excel_path)
    employee_rows = read_employee_sheet(excel_path)
    hour_rows = read_hours_sheet(excel_path)

    services = fetch_services(client, business["id"])
    staff_lookup = fetch_staff(client, business["id"])
    customers = fetch_customers(client, business["id"])
    appointment_markers = fetch_existing_markers(
        client,
        business["id"],
        "appointments",
        "internal_notes",
    )
    payment_markers = fetch_existing_markers(
        client,
        business["id"],
        "payments",
        "transaction_id",
    )
    expense_markers = fetch_existing_markers(
        client,
        business["id"],
        "expenses",
        "source",
    )
    payout_markers = fetch_existing_markers(
        client,
        business["id"],
        "payouts",
        "source",
    )
    time_log_markers = fetch_existing_markers(
        client,
        business["id"],
        "staff_time_logs",
        "source",
    )

    review_items: list[dict[str, Any]] = []
    summary = ImportSummary()

    sync_services(
        client,
        apply_changes=args.apply,
        business_id=business["id"],
        base_rows=base_rows,
        services=services,
        summary=summary,
    )
    sync_staff(
        client,
        apply_changes=args.apply,
        business_id=business["id"],
        employee_rows=employee_rows,
        staff_lookup=staff_lookup,
        review_items=review_items,
        summary=summary,
    )
    import_base_rows(
        client,
        apply_changes=args.apply,
        business_id=business["id"],
        base_rows=base_rows,
        services=services,
        customers=customers,
        staff_lookup=staff_lookup,
        appointment_markers=appointment_markers,
        payment_markers=payment_markers,
        expense_markers=expense_markers,
        payout_markers=payout_markers,
        review_items=review_items,
        summary=summary,
    )
    import_hour_rows(
        client,
        apply_changes=args.apply,
        business_id=business["id"],
        hour_rows=hour_rows,
        staff_lookup=staff_lookup,
        time_log_markers=time_log_markers,
        review_items=review_items,
        summary=summary,
    )
    flush_customer_updates(
        client,
        apply_changes=args.apply,
        customers=customers,
        summary=summary,
    )

    report = build_review_report(
        excel_path=excel_path,
        business_slug=business_slug,
        apply_changes=args.apply and not args.dry_run,
        summary=summary,
        review_items=review_items,
    )
    args.review_file.write_text(
        json.dumps(report, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    print(json.dumps(report["summary"], indent=2, ensure_ascii=False))
    print(f"Reporte guardado en {args.review_file}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
